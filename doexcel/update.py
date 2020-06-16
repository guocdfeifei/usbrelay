#author=guoff
#date  18:05

# basedir=r'D:\PycharmProjects\usbrelay\doexcel'
excelpath = r'三项岗位人员信息表.xlsx'

import xlrd
from xlutils.copy import copy
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


def excelUpdate(excel):
    '''
    更新已有excel, 是copy一份再做更新
    :param excel:
    :return:
    '''
    wb = xlrd.open_workbook(excel)#xlrd.open_workbook(excel,formatting_info=True)
    # newb = copy(wb)  # 类型为worksheet 无nrows 方法
    # wbsheet = newb.get_sheet(1)
    # # myWorkbook = xlrd.open_workbook(excel)
    mySheets = wb.sheets()  # 获取工作表list。
    #
    mySheet = mySheets[1]  # 通过索引顺序获取。
    #
    nrows = mySheet.nrows
    #
    # ncols = mySheet.ncols





    # wbsheet.write(19, 4, "abcd")
    #
    # wbsheet.write(20, 4, 1000)
    # newb.save(basedir+r"\newexcel.xlsx")

    #更新图片
    import glob
    # 创建一个新Excel文件并添加一个工作表。
    wb = load_workbook(r'D:\PycharmProjects\usbrelay\doexcel\三项岗位人员信息表.xlsx')
    worksheet = wb.worksheets[1]
    print('worksheet',worksheet)
    piclist = glob.glob(r"./*.jpg")
    print('piclist', piclist)
    for row in range(2, nrows):
        myCell = mySheet.cell(row, 1)  # 获取单元格，i是行数，j是列数，行数和列数都是从0开始计数。

        myCellValue = myCell.value
        print('myCellValue', myCellValue)
        for pic in piclist:
            if myCellValue in pic:
                print('pic',pic)
                img = Image(pic)
                img.width = 340
                img.height = 190
                worksheet.delete_cols(row, 3)
                worksheet.add_image(img, 'D'+str(row+1))
                # worksheet.column_dimensions['D'].width = 350
                # worksheet.row_dimensions[row+1].height = 200
                # worksheet.insert_image(row, 3, pic)

    wb.save(r'./三项岗位人员信息表1.xlsx')
#
# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image
#
#
#     excel_address = r"E:\Coding\E_PythonWriting\Excel\openpyxl示例_6.xlsx"
#     wb = load_workbook(excel_address)
#     sht = wb.worksheets[0]
#
#     img_address_1 = r"E:\Coding\E_PythonWriting\Excel\1.png"
#     img = Image(img_address_1)
#     sht.add_image(img, 'A1')
#
#     sht.column_dimensions['K'].width = 20.0
#     sht.row_dimensions[1].height = 40.0
#
#     img_address_2 = r"E:\Coding\E_PythonWriting\Excel\1.png"
#     img = Image(img_address_2)
#     img.width = 19.0
#     img.height = 39.0
#
#     sht.add_image(img, 'K1')
#
#     wb.save(excel_address)
#

if __name__ == "__main__":
    excelUpdate(excelpath)